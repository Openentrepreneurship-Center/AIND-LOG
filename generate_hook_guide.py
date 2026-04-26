#!/usr/bin/env python3
"""
Cline Hook 8종 설명 + 실제 수집 데이터 기반 정리 엑셀 생성
출력: <프로젝트 루트>/cline_hook_guide.xlsx
"""
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT        = Path(__file__).resolve().parent
EVENTS_PATH = ROOT / ".cline-metrics" / "events.jsonl"
OUT_PATH    = ROOT / "cline_hook_guide.xlsx"

KST = timezone(timedelta(hours=9))

# ── 색상 ────────────────────────────────────────────────────────────────────
NAVY      = "1F4E78"
DARK_NAVY = "162E46"
WHITE     = "FFFFFF"
LIGHT_BG  = "EBF2FA"
GRAY_ROW  = "F7F9FC"

HOOK_COLORS = {
    "TaskStart":        ("1D4ED8", "DBEAFE"),   # blue
    "TaskResume":       ("92400E", "FEF3C7"),   # amber
    "TaskCancel":       ("991B1B", "FEE2E2"),   # red
    "TaskComplete":     ("065F46", "D1FAE5"),   # green
    "PreToolUse":       ("7C3AED", "EDE9FE"),   # violet
    "PostToolUse":      ("0F766E", "CCFBF1"),   # teal
    "UserPromptSubmit": ("6D28D9", "F3E8FF"),   # purple
    "PreCompact":       ("374151", "F3F4F6"),   # gray
}

# ── 실제 데이터 로드 ─────────────────────────────────────────────────────────
def load_events():
    if not EVENTS_PATH.exists():
        return []
    items = []
    with EVENTS_PATH.open("r", encoding="utf-8", errors="ignore") as f:
        for raw in f:
            line = raw.strip()
            if line:
                try:
                    items.append(json.loads(line))
                except:
                    pass
    return items

def ts_kst(ts_ms):
    if ts_ms is None:
        return ""
    try:
        dt = datetime.fromtimestamp(int(ts_ms) / 1000, tz=KST)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return str(ts_ms)

# ── 스타일 헬퍼 ──────────────────────────────────────────────────────────────
def font(bold=False, size=10, color="000000", italic=False):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_all():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value, bold=False, size=10, fg="000000",
             bg=None, h="left", v="top", wrap=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(bold=bold, size=size, color=fg, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    c.border    = border_all()
    if bg:
        c.fill = fill(bg)
    return c

# ── 훅 정의 데이터 ────────────────────────────────────────────────────────────
HOOK_DEFINITIONS = [
    {
        "name": "TaskStart",
        "when": "새 Task를 시작할 때",
        "trigger": "사용자가 Cline에 새 작업을 요청(프롬프트 입력 후 Send)",
        "purpose": "Task 시작 시점을 기록하여 전체 소요시간 측정의 기준점으로 사용. 초기 요청(spec) 내용도 함께 수집",
        "payload_keys": "taskMetadata (taskId, ulid, initialTask)",
        "key_data": "initialTask: 사용자의 초기 요청 텍스트",
        "metrics_use": "• 코드 생성 소요시간 측정 시작점\n• Task별 초기 요청(spec) 저장\n• 전체 Task 수 집계",
        "intuitive": "에이전트가 '지금부터 시작합니다'를 선언하는 순간.\n모든 측정의 0초 기준점이 되며, 어떤 작업을 요청받았는지 기록한다.",
        "code_def": "hookName = 'TaskStart'\npayload.taskMetadata.taskId  → 고유 작업 식별자\npayload.taskMetadata.initialTask → 초기 요청 텍스트\n타임스탬프(ts) → 소요시간 계산 기준점 (t=0)",
        "simple": "🚀 '작업 시작'\n\n사용자가 AI에게 '이거 만들어줘'라고 하면 바로 이 훅이 발동합니다.\n'언제 시작했는지', '뭘 요청했는지'를 기록합니다.",
        "collected": True,
        "count": 0,
        "examples": [],
    },
    {
        "name": "TaskResume",
        "when": "중단된 Task를 재개할 때",
        "trigger": "취소/중단된 Task를 Resume 버튼으로 다시 시작",
        "purpose": "Task가 중간에 끊기고 재개된 횟수를 추적. '재업무율' 산출의 분자 역할",
        "payload_keys": "taskMetadata, previousState (lastMessageTs, messageCount, conversationHistoryDeleted)",
        "key_data": "messageCount: 재개 전까지 쌓인 대화 수\nlastMessageTs: 이전 마지막 메시지 시각",
        "metrics_use": "• 재업무율(Rework Rate) = Resume 횟수 / Task 시작 횟수\n• 이전 대화 컨텍스트 규모 파악",
        "intuitive": "중단됐던 작업을 이어받는 순간.\n이전 대화가 몇 개 있었는지, 언제 끊겼는지를 함께 전달하여\n'얼마나 자주 중간에 멈추고 다시 했는지' 추적한다.",
        "code_def": "hookName = 'TaskResume'\npayload.previousState.messageCount → 재개 전 대화 누적 수\npayload.previousState.lastMessageTs → 중단 직전 마지막 메시지 타임스탬프\npayload.previousState.conversationHistoryDeleted → 히스토리 삭제 여부",
        "simple": "🔄 '다시 시작'\n\n작업 도중 취소했다가 다시 이어서 하는 경우입니다.\n이게 많을수록 AI에게 여러 번 수정을 요청했다는 의미입니다.",
        "collected": True,
        "count": 0,
        "examples": [],
    },
    {
        "name": "TaskCancel",
        "when": "실행 중인 Task를 취소할 때",
        "trigger": "사용자가 Cline 실행 중 Stop/Cancel 버튼 클릭",
        "purpose": "중단된 Task를 식별하고 completionStatus를 기록. TaskResume과 쌍으로 재업무 패턴 분석",
        "payload_keys": "taskMetadata (taskId, ulid, completionStatus: 'cancelled')",
        "key_data": "completionStatus: 'cancelled' 고정값",
        "metrics_use": "• Task 실패/중단 횟수 집계\n• 취소 후 Resume 패턴 파악\n• 최종 상태 결정(취소됨)",
        "intuitive": "에이전트가 하던 일을 사용자가 강제로 멈춘 순간.\n이후 Resume이 따라오면 '재업무', 그냥 끝나면 '중단 처리'로 분류된다.",
        "code_def": "hookName = 'TaskCancel'\npayload.taskMetadata.completionStatus = 'cancelled'\n→ TaskResume 없이 종료 시: 최종 상태 = '취소됨'\n→ TaskResume 이후 발생 시: 중간 취소 카운트로 집계",
        "simple": "⛔ '중단'\n\nAI가 작업하는 도중 사용자가 '그만'을 누른 경우입니다.\n몇 번 중단했는지 기록해서 작업 흐름이 얼마나 끊겼는지 볼 수 있습니다.",
        "collected": True,
        "count": 0,
        "examples": [],
    },
    {
        "name": "TaskComplete",
        "when": "Task가 정상 완료될 때",
        "trigger": "Cline이 작업을 마치고 완료 메시지를 전송",
        "purpose": "Task 완료 시점과 최종 결과물 요약을 기록. 코드 생성 완료 기준점으로 사용",
        "payload_keys": "taskMetadata (taskId, ulid, result)",
        "key_data": "result: 모델이 생성한 완료 결과 요약 텍스트",
        "metrics_use": "• 최종 코드 생성까지 총 소요시간 측정 종료점\n• 완료 Task 비율 산출\n• 결과물 요약 보존",
        "intuitive": "에이전트가 '다 됐습니다'를 선언하는 순간.\n완료 결과 요약 텍스트가 포함되며, TaskStart와의 차이로 전체 소요시간이 계산된다.",
        "code_def": "hookName = 'TaskComplete'\npayload.taskMetadata.result → 모델이 작성한 완료 요약 텍스트\ntime_to_complete = TaskComplete.ts - TaskStart.ts\n→ 전체 작업 소요시간(ms) 계산에 사용",
        "simple": "✅ '완료'\n\nAI가 요청한 작업을 모두 끝냈을 때 기록됩니다.\n'무엇을 만들었는지' 요약 내용과 '얼마나 걸렸는지'를 알 수 있습니다.",
        "collected": True,
        "count": 0,
        "examples": [],
    },
    {
        "name": "PreToolUse",
        "when": "Cline이 도구를 실행하기 직전",
        "trigger": "write_to_file / read_file / execute_command 등 도구 호출 직전 자동 발화",
        "purpose": "도구 실행 전 파라미터(파일 경로, 코드 내용, 커맨드 등)를 사전 캡처. 실행 취소/변경 가능 시점",
        "payload_keys": "toolName, parameters (absolutePath/path, content, command, requires_approval 등)",
        "key_data": "toolName: 사용 도구명\ncontent: 파일에 쓸 전체 소스 코드\ncommand: 실행할 터미널 명령",
        "metrics_use": "• 코드 생성 시점 탐지 (write_to_file 감지)\n• 소스 코드 변경 내용 사전 캡처\n• 도구 호출 의도 파악",
        "intuitive": "에이전트가 파일을 쓰거나 명령을 실행하기 '직전' 찍히는 스냅샷.\n실행 전이라 아직 결과가 없고, '무엇을 하려는지' 의도(파라미터)만 담긴다.\n코드가 실제로 파일에 쓰이기 직전 원본 코드를 캡처할 수 있는 유일한 시점.",
        "code_def": "hookName = 'PreToolUse'\npayload.toolName → 사용할 도구 이름\npayload.parameters.content → 기록될 소스 코드 전문\npayload.parameters.command → 실행할 터미널 커맨드\npayload.parameters.absolutePath → 대상 파일 절대 경로\n※ success / executionTimeMs 없음 (실행 전이므로)",
        "simple": "🔍 '도구 사용 예고'\n\nAI가 파일을 저장하거나 명령어를 실행하기 직전에 기록됩니다.\n'어떤 코드를 저장하려는지'를 미리 볼 수 있습니다.",
        "collected": True,
        "count": 0,
        "examples": [],
    },
    {
        "name": "PostToolUse",
        "when": "도구 실행이 완료된 직후",
        "trigger": "모든 도구 실행이 끝난 직후 자동 발화",
        "purpose": "도구 실행 결과(성공/실패), 소요시간, 실행 결과물을 기록. 핵심 지표 대부분이 이 이벤트에서 산출",
        "payload_keys": "toolName, parameters, result, success (true/false), executionTimeMs",
        "key_data": "executionTimeMs: 도구 실행 소요시간(ms)\nsuccess: 성공 여부\nresult: 실행 결과 텍스트",
        "metrics_use": "• 코드 생성 소요시간 측정\n• 테스트 실행 횟수 및 소요시간\n• 도구 성공률 산출\n• 실제 소스 코드 스냅샷 저장",
        "intuitive": "에이전트가 도구를 실제로 실행하고 난 '결과 리포트'.\nPreToolUse와 쌍을 이루며, 성공 여부·소요시간·실행 결과가 추가된다.\n대부분의 핵심 지표(코드 생성 시간, 테스트 횟수, 도구 성공률)가 여기서 나온다.",
        "code_def": "hookName = 'PostToolUse'\npayload.success = true | false\npayload.executionTimeMs → 도구 실행 소요시간(ms)\npayload.result → 실행 결과물 텍스트(파일 내용 / 터미널 출력)\ntime_to_first_code = 첫 write_to_file PostToolUse.ts - TaskStart.ts",
        "simple": "📋 '도구 사용 결과'\n\n파일 저장이나 명령어 실행이 끝났을 때 결과가 기록됩니다.\n'성공했는지', '얼마나 걸렸는지'를 알 수 있습니다.\n대부분의 핵심 수치가 이 기록에서 나옵니다.",
        "collected": True,
        "count": 0,
        "examples": [],
    },
    {
        "name": "UserPromptSubmit",
        "when": "사용자가 Cline에 메시지를 전송할 때",
        "trigger": "채팅창에서 Enter / Send 버튼 클릭",
        "purpose": "사용자가 보낸 모든 프롬프트를 기록. 재업무 시 추가 지시 내용 추적 및 요청 패턴 분석",
        "payload_keys": "prompt (사용자 입력 텍스트)",
        "key_data": "prompt: 사용자가 입력한 텍스트 전체",
        "metrics_use": "• 초기 요청 vs 재업무 프롬프트 구분\n• 재업무 시 사용자 피드백 내용 보존\n• 프롬프트 횟수 집계",
        "intuitive": "사용자가 채팅창에 입력하고 전송한 순간을 기록하는 훅.\n첫 번째 프롬프트는 '초기 요청', 이후는 '추가 지시'로 분류 가능하다.\nTaskResume 직후 오는 것은 재업무 지시 내용으로 특별 관리한다.",
        "code_def": "hookName = 'UserPromptSubmit'\npayload.prompt → 사용자 입력 텍스트 전문\n첫 번째 UserPromptSubmit → first_prompt (초기 요청)\nTaskResume 직후 UserPromptSubmit → resume_prompt (재업무 지시)\n빈 문자열('')도 기록됨 (Resume 후 별도 입력 없는 경우)",
        "simple": "💬 '사용자 메시지'\n\n사람이 AI에게 말을 걸 때마다 기록됩니다.\n처음 요청한 내용, 중간에 추가 지시한 내용을 모두 시간 순서대로 볼 수 있습니다.",
        "collected": True,
        "count": 0,
        "examples": [],
    },
    {
        "name": "PreCompact",
        "when": "Cline이 대화 히스토리를 압축하기 직전",
        "trigger": "컨텍스트 창이 가득 차 자동으로 히스토리 요약/삭제 직전 발화",
        "purpose": "컨텍스트 압축 시점을 기록. 압축으로 인한 정보 손실 추적 및 장기 Task 모니터링",
        "payload_keys": "(압축 전 상태 정보 - Cline 버전에 따라 상이)",
        "key_data": "이 프로젝트에서는 발생하지 않음 (단기 Task였기 때문)",
        "metrics_use": "• 컨텍스트 압축 빈도 추적\n• 장기 Task 여부 판단 지표\n• 압축 전 주요 내용 사전 백업",
        "intuitive": "AI의 '기억 한도'가 차기 직전 발동하는 경고 훅.\n대화가 너무 길어져서 오래된 내용을 지워야 할 때 발생한다.\n이 훅이 자주 뜨면 하나의 작업이 매우 길고 복잡하다는 신호다.",
        "code_def": "hookName = 'PreCompact'\n→ 이 프로젝트 미발생 (단기 Task, 컨텍스트 압축 불필요)\n발생 시: 압축 직전 대화 상태를 스냅샷으로 저장 가능\n장기 Task 탐지 기준: PreCompact 발생 횟수 > 0",
        "simple": "🗜️ '대화 기억 압축'\n\nAI와 대화가 너무 길어지면 오래된 내용을 요약하고 지웁니다.\n그 직전에 이 기록이 남습니다. 이 프로젝트에서는 작업이 짧아서 발생하지 않았습니다.",
        "collected": False,
        "count": 0,
        "examples": [],
    },
]

# 실제 데이터 집계
events = load_events()
from collections import defaultdict
by_type = defaultdict(list)
for ev in events:
    by_type[ev.get("event","")].append(ev)

for hook in HOOK_DEFINITIONS:
    name = hook["name"]
    evs  = by_type.get(name, [])
    hook["count"] = len(evs)
    hook["collected"] = len(evs) > 0

    for ev in evs[:3]:
        p   = ev.get("payload") or {}
        ts  = ts_kst(ev.get("ts"))
        model = ev.get("model") or {}
        model_str = f"{model.get('provider','')}/{model.get('slug','')}" if model else ""

        if name == "TaskStart":
            meta = p.get("taskMetadata") or {}
            ex = f"ts: {ts}\ntaskId: {meta.get('taskId','')[:16]}\ninitialTask: {meta.get('initialTask','') or p.get('task','')}"
        elif name == "TaskResume":
            prev = p.get("previousState") or {}
            ex = f"ts: {ts}\nmessageCount: {prev.get('messageCount','')}\nlastMessageTs: {ts_kst(prev.get('lastMessageTs'))}\nconvHistoryDeleted: {prev.get('conversationHistoryDeleted','')}"
        elif name == "TaskCancel":
            meta = p.get("taskMetadata") or {}
            ex = f"ts: {ts}\ncompletionStatus: {meta.get('completionStatus','')}\ntaskId: {meta.get('taskId','')[:16]}"
        elif name == "TaskComplete":
            meta = p.get("taskMetadata") or {}
            result = str(meta.get("result",""))[:120]
            ex = f"ts: {ts}\nresult(앞120자): {result}"
        elif name in ("PreToolUse", "PostToolUse"):
            params = p.get("parameters") or {}
            tool   = p.get("toolName","")
            dur    = p.get("executionTimeMs","")
            succ   = p.get("success","")
            cmd    = str(params.get("command",""))[:80]
            path   = params.get("absolutePath","") or params.get("path","")
            content_len = len(params.get("content",""))
            ex = f"ts: {ts}\ntool: {tool}"
            if path: ex += f"\npath: {path}"
            if cmd:  ex += f"\ncommand: {cmd}"
            if content_len: ex += f"\ncontent 길이: {content_len}자"
            if dur:  ex += f"\nexecutionTimeMs: {dur} ({round(int(dur)/1000,1)}초)"
            if succ != "": ex += f"\nsuccess: {succ}"
        elif name == "UserPromptSubmit":
            ex = f"ts: {ts}\nprompt: {p.get('prompt','(빈 문자열 - TaskResume 직후)')}"
        else:
            ex = f"ts: {ts}\npayload: {json.dumps(p, ensure_ascii=False)[:200]}"

        if model_str:
            ex += f"\nmodel: {model_str}"
        hook["examples"].append(ex)

# ── 워크북 생성 ───────────────────────────────────────────────────────────────
wb  = Workbook()

# ════════════════════════════════════════════════════════════════════
# Sheet 1: 훅 개요 (한눈에 보는 요약표)
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "훅_개요"

# 타이틀
ws1.merge_cells("A1:M1")
c = ws1.cell(1, 1, "Cline Hook 8종 개요 – 실제 수집 데이터 기반 정리")
c.font = font(bold=True, size=14, color=WHITE)
c.fill = fill(NAVY)
c.alignment = align(h="center", v="center")
ws1.row_dimensions[1].height = 32

ws1.merge_cells("A2:M2")
c = ws1.cell(2, 1, f"프로젝트: clinetest 2 (계산기 앱)  |  총 수집 이벤트: {len(events)}건  |  생성일: {datetime.now(KST).strftime('%Y-%m-%d %H:%M')}")
c.font = font(italic=True, size=10, color="4B5563")
c.alignment = align(h="center", v="center")
ws1.row_dimensions[2].height = 18

ws1.append([])

HEADERS_1 = [
    "훅 이름", "실행 시점", "트리거 조건",
    "수집 여부", "이 프로젝트\n수집 건수",
    "수집 payload 키", "핵심 데이터",
    "지표 활용 목적",
    "직관적 설명 (기술)",
    "코드 관점 정의",
    "한눈에 보는 설명\n(비개발자용)",
]
header_row = 4
for col, h in enumerate(HEADERS_1, 1):
    set_cell(ws1, header_row, col, h, bold=True, size=10, fg=WHITE, bg=NAVY, h="center", v="center")
ws1.row_dimensions[header_row].height = 30

for row_idx, hook in enumerate(HOOK_DEFINITIONS, start=5):
    name = hook["name"]
    color_fg, color_bg = HOOK_COLORS.get(name, ("374151", "F3F4F6"))
    bg = GRAY_ROW if row_idx % 2 == 0 else "FFFFFF"

    collected_str = "✅ 수집됨" if hook["collected"] else "⬜ 미발생"
    collected_fg  = "065F46" if hook["collected"] else "6B7280"

    data = [
        hook["name"],
        hook["when"],
        hook["trigger"],
        collected_str,
        str(hook["count"]) + "건",
        hook["payload_keys"],
        hook["key_data"],
        hook["metrics_use"],
        hook.get("intuitive", ""),
        hook.get("code_def", ""),
        hook.get("simple", ""),
    ]
    for col, val in enumerate(data, 1):
        c = set_cell(ws1, row_idx, col, val, size=10, bg=bg)
        if col == 1:
            c.font = font(bold=True, size=10, color=color_fg)
            c.fill = fill(color_bg)
        if col == 4:
            c.font = font(bold=True, size=10, color=collected_fg)
        if col == 9:   # 직관적 설명 - 이탤릭
            c.font = font(size=10, italic=True, color="1E3A5F")
        if col == 10:  # 코드 관점 정의 - 고정폭 느낌
            c.font = Font(name="Consolas", size=9, color="374151")
            c.alignment = align(h="left", v="top", wrap=True)
        if col == 11:  # 비개발자용 - 약간 큰 폰트
            c.font = font(size=10, color="065F46")

    ws1.row_dimensions[row_idx].height = 90

col_widths_1 = [18, 22, 30, 12, 10, 35, 35, 40, 42, 45, 42]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
# Sheet 2: 훅별 상세 (payload 구조 + 실제 데이터 예시)
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("훅별_상세")

ws2.merge_cells("A1:F1")
c = ws2.cell(1, 1, "Cline Hook 8종 – payload 구조 및 실제 수집 데이터 예시")
c.font = font(bold=True, size=14, color=WHITE)
c.fill = fill(NAVY)
c.alignment = align(h="center", v="center")
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A2:F2")
c = ws2.cell(2, 1, "각 훅이 실제로 어떤 데이터를 담아 전달하는지 이 프로젝트(계산기 앱)의 실제 수집 로그 기반으로 정리")
c.font = font(italic=True, size=10, color="4B5563")
c.alignment = align(h="center", v="center")
ws2.row_dimensions[2].height = 18

ws2.append([])

HEADERS_2 = ["훅 이름", "실행 시점 & 목적", "payload 전체 구조", "이 프로젝트 실제 데이터 예시 (1)", "실제 데이터 예시 (2)", "실제 데이터 예시 (3) / 미발생 사유"]
for col, h in enumerate(HEADERS_2, 1):
    set_cell(ws2, 4, col, h, bold=True, size=10, fg=WHITE, bg=NAVY, h="center", v="center")
ws2.row_dimensions[4].height = 28

for row_idx, hook in enumerate(HOOK_DEFINITIONS, start=5):
    name = hook["name"]
    color_fg, color_bg = HOOK_COLORS.get(name, ("374151", "F3F4F6"))
    bg = GRAY_ROW if row_idx % 2 == 0 else "FFFFFF"

    purpose_text = f"{hook['when']}\n\n{hook['purpose']}"

    payload_struct = f"키 목록:\n{hook['payload_keys']}\n\n핵심 필드:\n{hook['key_data']}"

    ex1 = hook["examples"][0] if len(hook["examples"]) > 0 else ""
    ex2 = hook["examples"][1] if len(hook["examples"]) > 1 else ""
    ex3_label = hook["examples"][2] if len(hook["examples"]) > 2 else (
        "※ 이 프로젝트에서 미발생\n사유: 단기 Task(컨텍스트 압축 불필요)"
        if not hook["collected"] else "")

    row_data = [name, purpose_text, payload_struct, ex1, ex2, ex3_label]

    for col, val in enumerate(row_data, 1):
        c = set_cell(ws2, row_idx, col, val, size=9, bg=bg)
        if col == 1:
            c.font = font(bold=True, size=10, color=color_fg)
            c.fill = fill(color_bg)
            c.alignment = align(h="center", v="center")

    ws2.row_dimensions[row_idx].height = 100

col_widths_2 = [18, 28, 35, 40, 40, 40]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
# Sheet 3: 이벤트 타임라인 (실제 수집 이벤트 전체 흐름)
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("이벤트_타임라인")

ws3.merge_cells("A1:H1")
c = ws3.cell(1, 1, "실제 수집된 이벤트 타임라인 – Task 흐름 재현")
c.font = font(bold=True, size=14, color=WHITE)
c.fill = fill(NAVY)
c.alignment = align(h="center", v="center")
ws3.row_dimensions[1].height = 32

ws3.merge_cells("A2:H2")
c = ws3.cell(2, 1, f"이 프로젝트에서 실제로 수집된 {len(events)}개 이벤트의 전체 흐름. 한 행 = 한 이벤트 (시간순)")
c.font = font(italic=True, size=10, color="4B5563")
c.alignment = align(h="center", v="center")
ws3.row_dimensions[2].height = 18
ws3.append([])

HEADERS_3 = ["순번", "시각(KST)", "훅 이름", "도구명", "핵심 내용", "소요(초)", "성공 여부", "이 이벤트의 의미"]
for col, h in enumerate(HEADERS_3, 1):
    set_cell(ws3, 4, col, h, bold=True, size=10, fg=WHITE, bg=NAVY, h="center", v="center")
ws3.row_dimensions[4].height = 24

MEANING_MAP = {
    "TaskStart":        "새 Task 시작 – 초기 요청 캡처, 소요시간 측정 시작",
    "TaskResume":       "중단된 Task 재개 – 재업무율 분자 +1",
    "TaskCancel":       "Task 취소 – 사용자가 중간에 중단",
    "TaskComplete":     "Task 정상 완료 – 모델 결과물 요약 저장",
    "PreToolUse":       "도구 실행 직전 – 파라미터(코드/커맨드) 사전 캡처",
    "PostToolUse":      "도구 실행 완료 – 소요시간·성공여부·결과 기록",
    "UserPromptSubmit": "사용자 프롬프트 제출 – 요청 내용 기록",
    "PreCompact":       "컨텍스트 압축 직전 (이 프로젝트 미발생)",
    "GitCommit":        "git commit 완료 – diff + 전체 스냅샷 저장",
}

for idx, ev in enumerate(events, start=1):
    name    = ev.get("event", "")
    ts      = ts_kst(ev.get("ts"))
    payload = ev.get("payload") or {}
    model   = ev.get("model") or {}
    color_fg, color_bg = HOOK_COLORS.get(name, ("374151", "F3F4F6"))
    bg = GRAY_ROW if idx % 2 == 0 else "FFFFFF"

    tool_name = payload.get("toolName","") or payload.get("tool","")
    dur_raw   = payload.get("executionTimeMs")
    dur_sec   = round(int(dur_raw)/1000, 1) if dur_raw else ""
    success   = payload.get("success","")
    success_str = "✅" if success is True else ("❌" if success is False else "")

    # 핵심 내용
    params = payload.get("parameters") or {}
    if name == "TaskStart":
        meta = payload.get("taskMetadata") or {}
        main = meta.get("initialTask","") or payload.get("task","")
    elif name in ("TaskCancel","TaskComplete","TaskResume"):
        meta = payload.get("taskMetadata") or {}
        prev = payload.get("previousState") or {}
        if name == "TaskComplete":
            main = str(meta.get("result",""))[:100]
        elif name == "TaskResume":
            main = f"messageCount={prev.get('messageCount','')} / lastMsg={ts_kst(prev.get('lastMessageTs'))}"
        else:
            main = f"completionStatus={meta.get('completionStatus','')}"
    elif name in ("PreToolUse","PostToolUse"):
        path = params.get("absolutePath","") or params.get("path","")
        cmd  = str(params.get("command",""))[:60]
        clen = len(params.get("content",""))
        main = f"tool={tool_name}"
        if path: main += f" | path={path}"
        if cmd:  main += f" | cmd={cmd}"
        if clen: main += f" | content={clen}자"
        result = str(payload.get("result",""))[:60]
        if result and name == "PostToolUse": main += f" | result={result}"
    elif name == "UserPromptSubmit":
        main = payload.get("prompt","(빈 문자열)")
    elif name == "GitCommit":
        main = f"sha={ev.get('sha','')[:10]} | {ev.get('message','')[:60]}"
    else:
        main = json.dumps(payload, ensure_ascii=False)[:100]

    row_data = [idx, ts, name, tool_name, main, dur_sec, success_str, MEANING_MAP.get(name,"")]

    for col, val in enumerate(row_data, 1):
        c = set_cell(ws3, idx+4, col, str(val) if val != "" else "", size=9, bg=bg)
        if col == 3:
            c.font = font(bold=True, size=9, color=color_fg)
            c.fill = fill(color_bg)
        if col in (1,6):
            c.alignment = align(h="center", v="top", wrap=False)

    ws3.row_dimensions[idx+4].height = 45

col_widths_3 = [8, 20, 18, 16, 55, 10, 10, 40]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
# Sheet 4: 지표 매핑 (훅 → 지표 관계)
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("지표_매핑")

ws4.merge_cells("A1:F1")
c = ws4.cell(1, 1, "수집 지표 ↔ 훅 매핑 – 각 지표가 어느 훅에서 산출되는가")
c.font = font(bold=True, size=14, color=WHITE)
c.fill = fill(NAVY)
c.alignment = align(h="center", v="center")
ws4.row_dimensions[1].height = 32

ws4.merge_cells("A2:F2")
c = ws4.cell(2, 1, "지표별 데이터 출처 훅, 산출 방법, 이 프로젝트 실측값 정리")
c.font = font(italic=True, size=10, color="4B5563")
c.alignment = align(h="center", v="center")
ws4.row_dimensions[2].height = 18
ws4.append([])

HEADERS_4 = ["지표명", "출처 훅(들)", "산출 방법", "이 프로젝트 실측값", "수집 가능 여부", "비고"]
for col, h in enumerate(HEADERS_4, 1):
    set_cell(ws4, 4, col, h, bold=True, size=10, fg=WHITE, bg=NAVY, h="center", v="center")
ws4.row_dimensions[4].height = 24

METRICS = [
    ("코드 생성 소요시간",
     "TaskStart\n+ PostToolUse",
     "TaskStart ts → 첫 write_to_file PostToolUse ts 까지의 차이",
     "39.7초 (1776070955478 → 1776070995165)",
     "✅ 수집 가능",
     "Dev spec 유무 구분 위해 프롬프트에 [SPEC] 태그 도입 권장"),
    ("전체 Task 소요시간",
     "TaskStart\n+ TaskComplete/Cancel",
     "TaskStart ts → TaskComplete/Cancel ts 까지의 차이",
     "4608.5초 (약 76.8분, TaskCancel 기준)",
     "✅ 수집 가능",
     "Resume 포함 전체 시간. 실제 작업 시간 ≠ 경과 시간"),
    ("단위 테스트 횟수",
     "PostToolUse",
     "execute_command 중 pytest/jest/npm test 등 정규식 패턴 매칭",
     "0건 (이 프로젝트: 테스트 없는 HTML 앱)",
     "✅ 구조 준비됨\n⬜ 실측값 없음",
     "테스트 프레임워크 사용 프로젝트에서만 수집 가능"),
    ("테스트 소요시간 비중",
     "PostToolUse\n+ TaskStart",
     "테스트 executionTimeMs 합계 / 전체 Task 소요시간 × 100",
     "0% (테스트 없음)",
     "✅ 구조 준비됨\n⬜ 실측값 없음",
     "전체 소요시간 중 테스트 비중 파악"),
    ("재업무율 (Rework Rate)",
     "TaskStart\n+ TaskResume",
     "TaskResume 건수 / TaskStart 건수 × 100",
     "100% (Task 1건 중 2번 Resume)",
     "✅ 수집 가능",
     "높을수록 중간 수정/재지시 많음"),
    ("재업무 프롬프트 내용",
     "TaskResume\n+ UserPromptSubmit",
     "TaskResume 직후 UserPromptSubmit payload.prompt",
     "'작업내역 커밋해줘'\n(빈 문자열 1건 포함)",
     "✅ 수집 가능",
     "재개 시 사용자가 추가로 내린 지시 파악"),
    ("커밋된 소스코드",
     "GitCommit\n(post-commit 훅)",
     "git commit → .patch 파일 + .snapshot.json 저장",
     "(예: 마지막 GitCommit 1건)",
     "✅ 수집 가능",
     "diff + 전체 스냅샷 동시 저장"),
    ("검수 완료 커밋",
     "GitCommit\n(post-commit 훅)",
     "커밋 메시지에 [reviewed] 포함 여부 → final/<sha>.marker",
     "0건 (아직 검수 태그 없음)",
     "✅ 구조 준비됨",
     "커밋 메시지에 [reviewed] 태그 붙이면 자동 수집"),
    ("도구별 사용 횟수",
     "PreToolUse\n+ PostToolUse",
     "toolName 기준 집계",
     "write_to_file: 2회\nread_file: 1회\nexecute_command: 1회",
     "✅ 수집 가능",
     "모델의 코드 작성 vs 읽기 vs 실행 비율 파악"),
    ("모델 정보",
     "모든 훅",
     "각 이벤트의 model.provider + model.slug",
     "cline / anthropic/claude-opus-4.6",
     "✅ 수집 가능",
     "멀티 모델 사용 시 모델별 성능 비교 가능"),
]

for row_idx, (metric, hooks, method, actual, available, note) in enumerate(METRICS, start=5):
    bg = GRAY_ROW if row_idx % 2 == 0 else "FFFFFF"
    avail_fg = "065F46" if "✅" in available else "6B7280"
    for col, val in enumerate([metric, hooks, method, actual, available, note], 1):
        c = set_cell(ws4, row_idx, col, val, size=9, bg=bg)
        if col == 1: c.font = font(bold=True, size=10)
        if col == 5: c.font = font(bold=True, size=9, color=avail_fg)
    ws4.row_dimensions[row_idx].height = 55

col_widths_4 = [22, 20, 40, 38, 18, 38]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A5"

# ── 저장 ─────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"생성 완료: {OUT_PATH}")
print("시트 목록:")
for ws in wb.worksheets:
    print(f"  - {ws.title}")
