#!/usr/bin/env python3
"""Export Cline metrics into a human-readable Excel workbook.

Sheets: Overview / Task_Summary / Event_Log / Raw_JSONL / Counts

Usage:
  python3 .cline-metrics/export_excel_friendly.py
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
EVENTS_PATH = ROOT / "events.jsonl"
FINAL_DIR = ROOT / "final"
OUT_DIR = ROOT / "excel_export"
XLSX_PATH = OUT_DIR / "cline_metrics.xlsx"

# ── constants ───────────────────────────────────────────────────────────────
KST = timezone(timedelta(hours=9))
TEST_CMD_RE = re.compile(
    r"\b(pytest|jest|vitest|mocha|go test|cargo test|npm test|yarn test|unittest)\b"
)
CODE_TOOLS = {"write_to_file", "replace_in_file", "new_file", "apply_diff"}

# styling
NAVY = "1F4E78"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
GRAY_ROW = "F5F5F5"


# ── helpers ─────────────────────────────────────────────────────────────────
def normalize_ts(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def ts_to_kst(ts_ms: int | None) -> str:
    if ts_ms is None:
        return ""
    dt = datetime.fromtimestamp(ts_ms / 1000, tz=KST)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def ms_to_min(ms: int | None) -> float | None:
    if ms is None:
        return None
    return round(ms / 60000, 2)


def ms_to_sec(ms: int | None) -> float | None:
    if ms is None:
        return None
    return round(ms / 1000, 3)


_ILLEGAL_CHARS_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f]"
)


def safe_str(v: Any, max_len: int = 500) -> str:
    if v is None:
        return ""
    s = _ILLEGAL_CHARS_RE.sub(" ", str(v))
    s = s[:max_len] + ("..." if len(s) > max_len else "")
    # 수식으로 오해받지 않도록 = + - @ 로 시작하는 값에 접두 공백 추가
    if s and s[0] in ("=", "+", "-", "@"):
        s = " " + s
    return s


# ── style helpers ────────────────────────────────────────────────────────────
def _navy_font(bold: bool = True) -> Font:
    return Font(name="맑은 고딕", bold=bold, color=WHITE, size=11)


def _white_bold(size: int = 14) -> Font:
    return Font(name="맑은 고딕", bold=True, color=WHITE, size=size)


def _navy_fill() -> PatternFill:
    return PatternFill("solid", fgColor=NAVY)


def _light_fill() -> PatternFill:
    return PatternFill("solid", fgColor=LIGHT_BLUE)


def _gray_fill() -> PatternFill:
    return PatternFill("solid", fgColor=GRAY_ROW)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _left_wrap() -> Alignment:
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def write_title(ws, title: str, subtitle: str) -> None:
    ws.append([title])
    ws.cell(ws.max_row, 1).font = _white_bold(14)
    ws.cell(ws.max_row, 1).fill = _navy_fill()
    ws.cell(ws.max_row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[ws.max_row].height = 26

    ws.append([subtitle])
    ws.cell(ws.max_row, 1).font = Font(name="맑은 고딕", italic=True, size=10, color="555555")
    ws.row_dimensions[ws.max_row].height = 16

    ws.append([])  # 빈 행


def write_header_row(ws, headers: list[str]) -> int:
    ws.append(headers)
    row_idx = ws.max_row
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row_idx, col)
        c.font = _navy_font(bold=True)
        c.fill = _navy_fill()
        c.alignment = _center()
    ws.row_dimensions[row_idx].height = 20
    return row_idx


def stripe_row(ws, row_idx: int, col_count: int) -> None:
    if row_idx % 2 == 0:
        fill = _gray_fill()
        for col in range(1, col_count + 1):
            ws.cell(row_idx, col).fill = fill


def auto_col_width(ws, min_w: int = 10, max_w: int = 50) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))


# ── load events ──────────────────────────────────────────────────────────────
def load_events() -> list[dict]:
    if not EVENTS_PATH.exists():
        return []
    items = []
    with EVENTS_PATH.open("r", encoding="utf-8", errors="ignore") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            try:
                items.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return items


# ── build data ───────────────────────────────────────────────────────────────
def build_data(events: list[dict]) -> dict:
    tasks: dict[str, dict] = defaultdict(lambda: {
        "start_ts": None, "end_ts": None,
        "first_prompt": "", "initial_task": "",
        "model_provider": "", "model_slug": "", "cline_version": "",
        "resumed": False, "canceled": False, "completed": False,
        "resume_count": 0, "cancel_count": 0, "complete_count": 0,
        "prompts": [], "tools_used": [], "tool_duration_ms": 0,
        "pre_tool_count": 0, "post_tool_count": 0,
        "success_tool_count": 0, "write_count": 0,
        "read_count": 0, "exec_count": 0,
        "file_paths": set(), "last_event": "", "last_result": "",
        "event_count": 0,
    })

    event_rows = []
    last_task_id = None

    for idx, ev in enumerate(events, start=1):
        tid = ev.get("taskId") or ""
        event = ev.get("event", "")
        ts = normalize_ts(ev.get("ts"))
        payload = ev.get("payload") or {}
        model = ev.get("model") or {}
        provider = model.get("provider", "")
        slug = model.get("slug", "")
        version = ev.get("clineVersion", "")

        # ── shared payload fields ──
        prompt = payload.get("prompt", "")
        initial_task = (payload.get("taskMetadata") or {}).get("initialTask") or payload.get("task", "")
        tool_name = payload.get("tool") or payload.get("toolName") or ""
        params = payload.get("parameters") or {}
        path_val = params.get("path", "")
        abs_path = params.get("absolutePath", "")
        command = params.get("command", "")
        requires_approval = params.get("requiresApproval", "")
        success = payload.get("success", "")
        _exec_raw = payload.get("durationMs") or payload.get("executionTimeMs")
        exec_time_sec = round(_exec_raw / 1000, 3) if isinstance(_exec_raw, (int, float)) else ""
        result_raw = payload.get("result") or ""
        result_preview = safe_str(result_raw, 300)
        git_sha = ev.get("sha", "")
        git_message = ev.get("message", "")
        content_preview = safe_str(params.get("content", ""), 300)
        combined_path = abs_path or path_val

        # progress info
        progress = payload.get("progress") or {}
        prog_text = payload.get("progressStatus") or progress.get("text", "")
        prog_done = progress.get("done", "")
        prog_total = progress.get("total", "")
        prog_pct = round(prog_done / prog_total * 100, 1) if prog_done and prog_total else ""

        # resume fields
        prev_ts = (payload.get("previousState") or {}).get("lastMessageTs", "")
        prev_msg_count = (payload.get("previousState") or {}).get("messageCount", "")
        prev_hist_del = (payload.get("previousState") or {}).get("historyDeleted", "")

        # boolean flags
        is_tool = 1 if event in ("PreToolUse", "PostToolUse") else 0
        is_write = 1 if tool_name == "write_to_file" else 0
        is_read = 1 if tool_name in ("read_file", "read_file_content") else 0
        is_exec = 1 if tool_name == "execute_command" else 0
        is_pretool = 1 if event == "PreToolUse" else 0
        is_posttool = 1 if event == "PostToolUse" else 0
        is_success = 1 if success is True else 0
        is_cancel = 1 if event == "TaskCancel" else 0
        is_complete = 1 if event == "TaskComplete" else 0
        is_resume = 1 if event == "TaskResume" else 0
        is_prompt = 1 if event == "UserPromptSubmit" else 0

        task_order = (tasks[tid]["event_count"] + 1) if tid else ""

        event_rows.append([
            idx, tid, event,
            str(ts) if ts else "",
            ts,
            ts_to_kst(ts),
            task_order,
            provider, slug, version,
            prompt, initial_task, initial_task,
            tool_name, path_val, abs_path,
            command, requires_approval, success, exec_time_sec,
            "", result_preview,
            git_sha, git_message,
            prev_ts, prev_msg_count, prev_hist_del,
            prog_text, prog_done, prog_total, prog_pct,
            content_preview,
            safe_str(json.dumps(payload, ensure_ascii=False), 500),
            is_tool, is_write, is_read, is_exec,
            is_pretool, is_posttool, is_success,
            is_cancel, is_complete, is_resume, is_prompt,
            combined_path,
        ])

        # ── update task aggregates ──
        if not tid:
            continue

        t = tasks[tid]
        t["event_count"] += 1
        t["last_event"] = event
        if provider:
            t["model_provider"] = provider
        if slug:
            t["model_slug"] = slug
        if version:
            t["cline_version"] = version

        if ts:
            if t["start_ts"] is None:
                t["start_ts"] = ts
            t["end_ts"] = ts

        if event == "TaskStart":
            t["initial_task"] = initial_task
            last_task_id = tid
        elif event == "UserPromptSubmit":
            if not t["first_prompt"]:
                t["first_prompt"] = prompt
            t["prompts"].append(prompt)
        elif event == "TaskResume":
            t["resumed"] = True
            t["resume_count"] += 1
        elif event == "TaskCancel":
            t["canceled"] = True
            t["cancel_count"] += 1
        elif event == "TaskComplete":
            t["completed"] = True
            t["complete_count"] += 1
            t["last_result"] = result_preview
        elif event == "PreToolUse":
            t["pre_tool_count"] += 1
            if tool_name and tool_name not in t["tools_used"]:
                t["tools_used"].append(tool_name)
        elif event == "PostToolUse":
            t["post_tool_count"] += 1
            if success is True:
                t["success_tool_count"] += 1
            if tool_name == "write_to_file":
                t["write_count"] += 1
            if tool_name in ("read_file", "read_file_content"):
                t["read_count"] += 1
            if tool_name == "execute_command":
                t["exec_count"] += 1
            if isinstance(_exec_raw, (int, float)):
                t["tool_duration_ms"] += _exec_raw
            if combined_path:
                t["file_paths"].add(combined_path)

    # ── task summary rows ──
    reviewed_shas = {p.stem for p in FINAL_DIR.glob("*.marker")} if FINAL_DIR.exists() else set()
    task_rows = []
    for tid, t in tasks.items():
        duration_ms = (t["end_ts"] - t["start_ts"]) if t["start_ts"] and t["end_ts"] else None
        if t["canceled"]:
            status = "취소됨"
        elif t["completed"]:
            status = "완료됨"
        elif t["resumed"]:
            status = "재개됨/진행중"
        else:
            status = "진행중/미확인"

        task_rows.append([
            tid,
            t["start_ts"],
            ts_to_kst(t["start_ts"]),
            t["end_ts"],
            ts_to_kst(t["end_ts"]),
            ms_to_min(duration_ms),
            t["event_count"],
            len(t["prompts"]),
            t["pre_tool_count"] + t["post_tool_count"],
            t["pre_tool_count"],
            t["post_tool_count"],
            t["success_tool_count"],
            t["write_count"],
            t["read_count"],
            t["exec_count"],
            t["resume_count"],
            t["cancel_count"],
            t["complete_count"],
            t["last_event"],
            status,
            t["first_prompt"],
            t["initial_task"],
            ", ".join(t["tools_used"]),
            ms_to_sec(t["tool_duration_ms"]) or 0,
            "\n".join(sorted(t["file_paths"]))[:300],
            t["last_result"],
            t["model_provider"],
            t["model_slug"],
            t["cline_version"],
        ])

    # ── counts ──
    event_type_counts = defaultdict(int)
    tool_counts = defaultdict(int)
    for ev in events:
        et = ev.get("event", "")
        event_type_counts[et] += 1
        if et in ("PreToolUse", "PostToolUse"):
            p = ev.get("payload") or {}
            tn = p.get("tool") or p.get("toolName") or ""
            if tn:
                tool_counts[tn] += 1

    total_events = len(events)
    count_rows = []
    for et, cnt in sorted(event_type_counts.items()):
        pct = f"{cnt/total_events*100:.1f}%" if total_events else ""
        count_rows.append(["이벤트종류", et, cnt, pct])
    for tn, cnt in sorted(tool_counts.items(), key=lambda x: -x[1]):
        total_tools = sum(tool_counts.values())
        pct = f"{cnt/total_tools*100:.1f}%" if total_tools else ""
        count_rows.append(["도구명", tn, cnt, pct])

    return {
        "events": events,
        "event_rows": event_rows,
        "task_rows": task_rows,
        "count_rows": count_rows,
        "reviewed_shas": reviewed_shas,
        "stats": {
            "total_events": total_events,
            "total_tasks": len(task_rows),
            "first_ts": min((normalize_ts(e.get("ts")) for e in events if e.get("ts")), default=None),
            "last_ts": max((normalize_ts(e.get("ts")) for e in events if e.get("ts")), default=None),
        },
    }


# ── sheet builders ───────────────────────────────────────────────────────────
def build_overview(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "Overview"
    stats = data["stats"]
    task_rows = data["task_rows"]
    event_rows = data["event_rows"]
    total = stats["total_events"]
    first_ts_str = ts_to_kst(stats["first_ts"])
    last_ts_str = ts_to_kst(stats["last_ts"])

    # 타이틀
    write_title(ws, "Cline 이벤트 로그 정리본",
                "events.jsonl 전체 데이터를 사람이 보기 좋게 정리한 엑셀입니다. 시간은 KST(한국시간) 기준입니다.")

    # 왼쪽 주요 지표
    def kv(label, value):
        ws.append([label, value])
        ws.cell(ws.max_row, 1).font = Font(name="맑은 고딕", bold=True, size=10)
        ws.cell(ws.max_row, 1).fill = _light_fill()
        ws.cell(ws.max_row, 2).font = Font(name="맑은 고딕", size=10)

    kv("총 이벤트 수", total)
    kv("총 작업(task) 수", stats["total_tasks"])
    kv("첫 이벤트 시각 (KST)", first_ts_str)
    kv("마지막 이벤트 시각 (KST)", last_ts_str)
    ws.append([])

    # 오른쪽 시트 안내 (별도 섹션)
    nav_title_row = ws.max_row + 1
    ws.append(["시트 안내"])
    ws.cell(ws.max_row, 1).font = Font(name="맑은 고딕", bold=True, color=WHITE, size=11)
    ws.cell(ws.max_row, 1).fill = _navy_fill()
    ws.row_dimensions[ws.max_row].height = 20

    for name, desc in [
        ("Overview", "핵심 지표와 시트 설명"),
        ("Task_Summary", "taskId 단위 요약 (소요시간·도구·상태 등)"),
        ("Event_Log", "한 줄 = 한 이벤트 (모든 필드 전개)"),
        ("Raw_JSONL", "원본 JSONL 전체 보관"),
        ("Counts", "이벤트 종류·도구별 개수와 비중"),
    ]:
        ws.append([name, desc])
        ws.cell(ws.max_row, 1).font = Font(name="맑은 고딕", bold=True, size=10)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55


def build_task_summary(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Task_Summary")
    write_title(ws, "작업(taskId) 단위 요약",
                "taskId별로 이벤트를 묶은 요약입니다. 핵심 컬럼을 한눈에 파악할 수 있도록 정리했습니다.")

    headers = [
        "taskId", "시작시각(epoch_ms)", "시작시각(KST)", "종료시각(epoch_ms)", "종료시각(KST)",
        "소요시간(분)", "이벤트수", "사용자프롬프트수", "도구이벤트수",
        "PreToolUse수", "PostToolUse수", "성공도구수",
        "write_to_file수", "read_file수", "execute_command수",
        "Resume수", "Cancel수", "Complete수",
        "마지막이벤트", "최종상태",
        "첫프롬프트", "초기요청",
        "사용도구목록", "총도구실행시간(초)", "주요파일경로",
        "마지막결과요약", "모델provider", "모델slug", "clineVersion",
    ]
    header_row = write_header_row(ws, headers)

    for i, row in enumerate(data["task_rows"], start=1):
        ws.append(row)
        r_idx = ws.max_row
        stripe_row(ws, r_idx, len(headers))
        for col in range(1, len(headers) + 1):
            ws.cell(r_idx, col).font = Font(name="맑은 고딕", size=10)
            ws.cell(r_idx, col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

    ws.freeze_panes = "A5"
    auto_col_width(ws, min_w=12, max_w=40)


def build_event_log(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Event_Log")
    write_title(ws, "이벤트 상세 로그",
                "원본 JSONL을 주요 컬럼으로 풀었습니다. 한 행 = 한 이벤트.")

    headers = [
        "line_no", "taskId", "event",
        "ts_raw", "ts_number", "ts_kst",
        "task내_순서",
        "model_provider", "model_slug", "clineVersion",
        "prompt", "initial_task", "task_name",
        "tool_name", "path", "absolute_path",
        "command", "requires_approval", "success", "실행시간(초)",
        "completion_status", "result_preview",
        "git_sha", "git_message",
        "prev_lastMessageTs", "prev_messageCount", "prev_history_deleted",
        "progress_text", "progress_done", "progress_total", "progress_pct",
        "content_preview", "raw_payload_preview",
        "is_tool_event", "is_write_to_file", "is_read_file", "is_execute_command",
        "is_pretool", "is_posttool", "is_success",
        "is_cancel", "is_complete", "is_resume", "is_prompt",
        "combined_path",
    ]
    write_header_row(ws, headers)

    for i, row in enumerate(data["event_rows"], start=1):
        ws.append(row)
        r_idx = ws.max_row
        stripe_row(ws, r_idx, len(headers))
        for col in range(1, len(headers) + 1):
            ws.cell(r_idx, col).font = Font(name="맑은 고딕", size=10)
            ws.cell(r_idx, col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

    ws.freeze_panes = "A5"
    auto_col_width(ws, min_w=10, max_w=45)


def build_raw_jsonl(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Raw_JSONL")
    write_title(ws, "원본 JSONL 보관",
                "Excel 셀 길이 제한 때문에 긴 JSON 줄은 part_1 / part_2 / part_3으로 분할했습니다.")

    headers = ["line_no", "part_1", "part_2", "part_3"]
    write_header_row(ws, headers)

    def _clean(s: str) -> str:
        s = _ILLEGAL_CHARS_RE.sub(" ", s)
        # 수식으로 오해받지 않도록 = + - @ 로 시작하는 셀에 접두 공백 추가
        if s and s[0] in ("=", "+", "-", "@"):
            s = " " + s
        return s

    for idx, ev in enumerate(data["events"], start=1):
        raw = json.dumps(ev, ensure_ascii=False)
        raw = _clean(raw)
        chunk = 30000
        parts = [raw[i:i+chunk] for i in range(0, len(raw), chunk)]
        ws.append([
            idx,
            parts[0] if len(parts) > 0 else "",
            parts[1] if len(parts) > 1 else "",
            parts[2] if len(parts) > 2 else "",
        ])
        r_idx = ws.max_row
        stripe_row(ws, r_idx, 4)
        for col in range(1, 5):
            ws.cell(r_idx, col).font = Font(name="맑은 고딕", size=9)
            # 명시적 문자열 타입으로 지정해 수식 오인 방지
            ws.cell(r_idx, col).data_type = "s"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 40
    ws.freeze_panes = "A5"


def build_counts(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Counts")
    write_title(ws, "분류별 집계",
                "이벤트 종류와 사용 도구 기준으로 개수와 비중을 집계했습니다.")

    headers = ["구분", "값", "개수", "비중"]
    write_header_row(ws, headers)

    for row in data["count_rows"]:
        ws.append(row)
        r_idx = ws.max_row
        stripe_row(ws, r_idx, 4)
        for col in range(1, 5):
            ws.cell(r_idx, col).font = Font(name="맑은 고딕", size=10)

    auto_col_width(ws, min_w=12, max_w=40)
    ws.freeze_panes = "A5"


# ── main ─────────────────────────────────────────────────────────────────────
def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    events = load_events()
    data = build_data(events)

    wb = Workbook()
    build_overview(wb, data)
    build_task_summary(wb, data)
    build_event_log(wb, data)
    build_raw_jsonl(wb, data)
    build_counts(wb, data)

    wb.save(XLSX_PATH)
    stats = data["stats"]
    print(f"완료: {XLSX_PATH}")
    print(f"  총 이벤트: {stats['total_events']}개")
    print(f"  총 작업:   {stats['total_tasks']}개")
    print(f"  Event_Log: {len(data['event_rows'])}행")
    print(f"  Counts:    {len(data['count_rows'])}행")


if __name__ == "__main__":
    main()
